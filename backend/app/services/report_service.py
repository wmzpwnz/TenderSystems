import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from typing import List, Dict, Any
from datetime import datetime

class ReportService:
    def generate_tenders_excel(self, tenders: List[Dict[str, Any]]) -> BytesIO:
        """
        Генерация Excel файла со списком тендеров.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Тендеры"

        # Заголовки
        headers = [
            "Номер", "Наименование", "Заказчик", "Начальная цена", 
            "Валюта", "Дата публикации", "Дедлайн", "Статус", "Ссылка"
        ]
        
        # Стилизация заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Преобразование данных
        for row_num, tender in enumerate(tenders, 2):
            # Извлекаем данные, учитывая возможные разные структуры (ORM vs Dict)
            if hasattr(tender, "__dict__"):
                t = tender.__dict__
            else:
                t = tender

            row_data = [
                t.get("number") or t.get("eis_id"),
                t.get("title"),
                t.get("customer_name"),
                t.get("initial_price"),
                t.get("currency", "RUB"),
                self._format_date(t.get("publication_date")),
                self._format_date(t.get("application_deadline")),
                t.get("status"),
                f"https://zakupki.gov.ru/epz/order/notice/view/common-info.html?regNumber={t.get('eis_id')}"
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                
                # Настройка переноса текста для длинных наименований
                if col_num in [2, 3]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(vertical="top")

        # Автоматическая ширина колонок
        column_widths = [20, 50, 40, 15, 10, 20, 20, 15, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Сохранение в буфер
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _format_date(self, dt):
        if not dt:
            return ""
        if isinstance(dt, datetime):
            return dt.strftime("%d.%m.%Y %H:%M")
        if isinstance(dt, str):
            try:
                # Пытаемся распарсить ISO формат
                parsed = datetime.fromisoformat(dt.replace('Z', '+00:00'))
                return parsed.strftime("%d.%m.%Y %H:%M")
            except:
                return dt
        return str(dt)

report_service = ReportService()
